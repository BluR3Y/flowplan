from __future__ import annotations
from typing import Dict, Optional, Sequence, List
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

import logging
import os
import pandas as pd

from .collection_manager import CollectionManager
from .sheet_manager import SheetManager

log = logging.getLogger(__name__)

class WorkbookManager(CollectionManager):
    """
    Orchestrates reading/writing Excel files.
    Inherits generic collection logic but specializes in SheetManagers
    """

    # Tell the parent class to create SheetManagers, not DocumentManagers
    _document_class = SheetManager

    def __init__(
        self,
        read_file_path: Optional[str] = None,
        *,
        include_sheets: Optional[Sequence[str]] = None
    ) -> None:
        super().__init__()
        self.read_file_path = read_file_path
        self.include_sheets = include_sheets

        self.write_file_path: Optional[str] = None
        self._wb: Optional[Workbook] = None
    
    # --------------------------- Helpers ---------------------------
    
    def create_sheet(self, name: str, data: SheetManager.DataType) -> SheetManager:
        return super().create_document(name, data)
    
    def get_sheet(self, name: str) -> SheetManager:
        return super().get_document(name)
    
    def delete_sheet(self, name: str) -> SheetManager:
        return super().delete_document(name)

    # --------------------------- I/O ---------------------------
    def open(self) -> None:
        """Loads the workbook if path exists, otherwise creates a new one."""
        if self.read_file_path and os.path.exists(self.read_file_path):
            log.info("Opening workbook: %s", self.read_file_path)
            # keep_vba=True is default in openpyxl, but good to know we are keeping it
            self._wb = load_workbook(self.read_file_path, data_only=False)
            self.load_existing_sheets(self.include_sheets)
        else:
            log.info("Creating new in-memory workbook.")
            self._wb = Workbook()
            self._dirty = True
    
    def close(self) -> None:
        if self._wb:
            self._wb.close()
            self._wb = None
    
    def load_existing_sheets(self, sheet_names: Optional[Sequence[str]] = None) -> None:
        """
        Loads sheets from the file into pandas DataFrames.
        NOTE: We read data using pandas for speed, separate from the openpyxl object.
        """
        if not self.read_file_path:
            return
        
        # Use pandas to read data (faster than iterating openpyxl)
        try:
            # Note: We must read 'data_only' implicitly via pandas
            # (pandas usually evaluates formulas or reads cached values)
            xls_file = pd.ExcelFile(self.read_file_path)
            all_names = xls_file.sheet_names

            target_names = sheet_names if sheet_names else all_names

            for name in target_names:
                if name in all_names:
                    df = pd.read_excel(xls_file, sheet_name=name)
                    # Register without triggering dirty flag
                    self._docs_created[name] = SheetManager(name, df, on_change=self.mark_dirty)
        except Exception as e:
            log.error("Failed to load sheets via pandas: %s", e)
            raise
    
    def set_write_path(self, path: str, *, allow_overwrite: bool = False) -> None:
        if not path:
            raise ValueError("Write path cannot be empty.")
        
        if (
            self.read_file_path and
            os.path.abspath(path) == os.path.abspath(self.read_file_path) and
            not allow_overwrite
        ):
            raise ValueError(
                f"Write path '{path}' matches read path."
                "Set allow_overwrite=True to enable destructive save."
            )
        self.write_file_path = path
    
    # --------------------------- Save Logic ---------------------------
    def save(self) -> Optional[str]:
        """
        Syncs modified DataFrames back to the OpenPyXL workbook and saves to disk.
        """
        if not self.write_file_path:
            log.warning("No write path set. Save aborted.")
            return None
        
        if not self._dirty:
            log.info("No changes detected. Save skipped.")
            return None
        
        if self._wb is None:
            self._wb = Workbook()

        try:
            # Handle Deletions: Remove sheets that were deleted via delete_document
            if self._docs_deleted:
                for name in list(self._docs_deleted):
                    if name in self._wb.sheetnames:
                        log.info(f"Deleting sheet '{name}' from workbook.")
                        del self._wb[name]
                # Clear the deletion queue after processing so subsequent saves are clean
                self._docs_deleted.clear()
            
            # Handle Updates/Creations
            for name, doc in self._docs_created.items():
                if isinstance(doc, SheetManager):
                    self._sync_sheet_to_workbook(name, doc)
                else:
                    log.warning(f"Document '{name}' is not a SheetManager. Skipping save.")
            
            self._wb.save(self.write_file_path)
            self._dirty = False
            return self.write_file_path
        except Exception as e:
            log.error("Save failed: %s", e)
            raise
    
    def _sync_sheet_to_workbook(self, name: str, sm: SheetManager) -> None:
        """
        Writes the DataFrame data into the OpenPyXL worksheet.
        Strategies:
        1. If sheet exists, we clear data (keeping generic sheet properties) and rewrite.
        2. If sheet missing, create new.
        """
        if name in self._wb.sheetnames:
            ws = self._wb[name]
            # Clear existing data to prevent zombie rows if new data is shorter
            # Note: delete_rows is slow for massive sheets, but safe for styles
            if ws.max_row > 0:
                ws.delete_rows(1, ws.max_row)
        else:
            ws = self._wb.create_sheet(title=name)
        
        # Prepare data
        df = sm.get_df()

        # Fast write using openpyxl utility
        # dataframe_to_rows is a generator; we iterate to append
        rows = dataframe_to_rows(df, index=False, header=True)

        for row in rows:
            ws.append(row)
        
        self._apply_annotations(ws, sm)
    
    def _apply_annotations(self, ws: Worksheet, sm: SheetManager) -> None:
        """Apply colors and comments."""
        fills = {
            "info":  PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
            "warn":  PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
            "error": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        }
        
        for (row, col, level, message) in sm.iter_annotations():
            # Adjust 0-based index to 1-based Excel index
            # +1 for 0->1 base, +1 for Header row
            erow = row + 2 
            ecol = col + 1 
            
            # Ensure cell exists (in case annotation is out of data bounds)
            if erow > ws.max_row:
                continue 
                
            cell = ws.cell(row=erow, column=ecol)
            
            if message:
                current_text = cell.comment.text + "\n" if cell.comment else ""
                cell.comment = Comment(current_text + message, "wbm")
                
            if level in fills:
                cell.fill = fills[level]
    
    # --------------------------- Context Manager ---------------------------
    def __enter__(self) -> "WorkbookManager":
        self.open()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        # Only attempt save if no error occurred
        if exc_type is None:
            self.save()
        self.close()