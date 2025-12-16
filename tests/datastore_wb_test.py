import pytest
import pandas as pd
from datastore.workbook_manager import WorkbookManager, SheetManager
from openpyxl import load_workbook

# 1. Use 'function' scope (default) for isolation. 
#    If one test deletes a row, others aren't affected.
@pytest.fixture
def manager(tmp_path):
    """
    Creates a fresh WorkbookManager for each test.
    Uses a temporary file path so we don't overwrite real files.
    """
    # Create a dummy file path in the temp directory
    fpath = tmp_path / "test_workbook.xlsx"
    
    # Initialize properly
    wm = WorkbookManager(str(fpath))
    
    # Manually trigger open() logic if not using 'with' block inside test
    wm.open()
    
    yield wm
    
    # Teardown: ensure we close/cleanup
    wm.close()

@pytest.fixture
def sheet_manager(manager):
    """
    Pre-populates the workbook with a standard dataset.
    """
    data = {
        "name": ["Alice", "Bob", "Charlie", "James"],
        "age": [22, 16, 31, 29],
        "color": ["Red", "Orange", "Blue", "Green"]
    }
    
    # Use the public API to create the document
    # Note: Ensure method name matches your implementation (create_document vs create_sheet)
    return manager.create_document("Sheet1", data)

def test_find_match(sheet_manager: SheetManager):
    """Test the matching logic on the pre-populated sheet."""
    # Act
    match = sheet_manager.match({"name": "Charlie"}, include_score=False)
    
    # Assert
    assert not match.empty
    match_data = match.to_dict(orient="records")[0]
    assert match_data == {'name': 'Charlie', 'age': 31, 'color': 'Blue'}

def test_add_annotation(sheet_manager: SheetManager):
    """Test that annotations can be added without error."""
    sheet_manager.add_annotation(0, "name", "info", "Check this")
    
    # Verify internal state
    annotations = list(sheet_manager.iter_annotations())
    assert len(annotations) == 1
    assert annotations[0] == (0, 0, "info", "Check this") # Assuming "name" is col 0

def test_save_logic(manager, tmp_path):
    """Test that saving actually writes a file."""
    # Setup
    manager.set_write_path(str(tmp_path / "output.xlsx"))
    manager.create_document("TestSheet", [{"col1": 1}])
    
    # Act
    saved_path = manager.save()
    
    # Assert
    assert saved_path is not None
    assert (tmp_path / "output.xlsx").exists()

def test_delete_document_logic(manager):
    """Test the internal state of CollectionManager after deletion."""
    # Setup
    manager.create_document("Sheet1", [{"col": 1}])
    manager.create_document("Sheet2", [{"col": 2}])
    
    assert "Sheet1" in manager.list_documents()
    
    # Act
    manager.delete_document("Sheet1")
    
    # Assert
    assert "Sheet1" not in manager.list_documents()
    assert "Sheet1" in manager._docs_deleted # Verify internal tracking
    assert "Sheet2" in manager.list_documents() # Ensure others are untouched

def test_delete_non_existent_document(manager):
    """Test that deleting a non-existent document raises KeyError."""
    with pytest.raises(KeyError):
        manager.delete_document("GhostSheet")

def test_recreation_after_deletion(manager, tmp_path):
    """
    CRITICAL EDGE CASE: 
    Test that deleting "A" and then creating a new "A" before saving
    results in the NEW "A" being saved, not the deletion of "A".
    """
    write_path = str(tmp_path / "recreate.xlsx")
    manager.set_write_path(write_path)
    
    # 1. Create Initial State
    manager.create_document("SheetA", [{"val": "Original"}])
    manager.save() 
    
    # 2. Delete and immediately Re-create
    manager.delete_document("SheetA")
    assert "SheetA" in manager._docs_deleted # Should be marked for death
    
    # 3. Re-create (The Un-delete logic)
    manager.create_document("SheetA", [{"val": "New Version"}])
    
    # Verify internal state BEFORE save
    assert "SheetA" not in manager._docs_deleted # Should be rescued
    assert "SheetA" in manager.list_documents()
    
    # 4. Save and Verify File
    manager.save()
    
    # Load raw excel to verify
    wb = load_workbook(write_path)
    assert "SheetA" in wb.sheetnames
    ws = wb["SheetA"]
    # Check that it contains the NEW data (header + row = 2 rows usually)
    # This assumes simple writing. We can check specific cell value if needed.
    # OpenPyXL cell access is [row][col] 1-based.
    # Header: val (A1), Data: New Version (A2)
    assert ws["A2"].value == "New Version"

def test_delete_persists_to_file(manager, tmp_path):
    """Test that saving actually removes the sheet from the .xlsx file."""
    write_path = str(tmp_path / "final_output.xlsx")
    manager.set_write_path(write_path)
    
    # 1. Setup a file with 2 sheets
    manager.create_document("KeepMe", [{"a": 1}])
    manager.create_document("DeleteMe", [{"b": 2}])
    manager.save()
    
    # Verify they exist
    wb_temp = load_workbook(write_path)
    assert "DeleteMe" in wb_temp.sheetnames
    wb_temp.close()
    
    # 2. Delete one
    manager.delete_document("DeleteMe")
    manager.save()
    
    # 3. Verify final file
    wb_final = load_workbook(write_path)
    assert "KeepMe" in wb_final.sheetnames
    assert "DeleteMe" not in wb_final.sheetnames